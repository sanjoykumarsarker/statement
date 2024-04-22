from django.shortcuts import render, redirect
from .forms import ImagesForm
from .models import Image
import glob
import os
import re
import openpyxl 
from pypdf import PdfReader     
file_list=[]
data_lines=[]
data_split_code=[]
data_split_account=[]
dir_list=[]
code_word=""
branch_name=""
account_word=""
path = "media/MiniApp_Images/"  

def index(request):
    dir_list=[] 
    for file in glob.glob(path + '*.pdf'):
        dir_list.append(file)
    return render(request, "index.html", {'file_list':dir_list})
def delete(request):
    dir_list=[] 
    files = glob.glob(path + '*.*')
    for f in files:
        os.remove(f)
    for file in glob.glob(path + '*.pdf'):
        dir_list.append(file)
    return render(request, "index.html", {'file_list':dir_list})
def process(request):
    dir_list=[] 
    for file in glob.glob(path + '*.pdf'):
        dir_list.append(file)
    r=1
    data=""
    wb = openpyxl.Workbook() 
    sheet = wb.active 
    for file in glob.glob(path + '*.pdf'):
        reader = PdfReader(file) 
        length=len(reader.pages)  
        for ln in range(length):
            page = reader.pages[ln] 
            data=" ".join([data, page.extract_text()])
 
    data_lines=data.splitlines()
    
    lngth=len(data_lines)
    for x in range(0, lngth):
        
        branch = re.findall("Branch Code", data_lines[x])
        code = re.findall("[0-9][0-9][0-9][0-9]\s*-", data_lines[x])
        account = re.findall("[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9]", data_lines[x])
        data_split_code.append(data_lines[x])
        if (branch):
            branch_name=data_lines[x]
        if (code):
            code_word=data_lines[x]
        if (account):
            account_word=data_lines[x]
            account_word_list=data_lines[x].split(" ")
            account_name_list=data_lines[x+2]
            account_balane=data_lines[x+5]
            
            
            c1 = sheet.cell(row = r, column = 1) 
            c1.value = branch_name[branch_name.find("Branch Code")+12:len(branch_name)]
            c1 = sheet.cell(row = r, column = 2) 
            c1.value = code_word 
            
            c2 = sheet.cell(row= r , column = 3) 
            c2.value = account_word_list[0]
            c2 = sheet.cell(row= r , column = 4) 
            c2.value = account_name_list
 
            
            c2 = sheet.cell(row= r , column = 5) 
            c2.value = account_word_list[2]
                        
            c2 = sheet.cell(row= r , column = 6) 
            c2.value = float(account_word_list[4].replace(",",""))

            #c2 = sheet.cell(row= r , column = 7) 
            #c2.value = account_balane

            r=r+1
        
    wb.save("media/MiniApp_Images/branch.xlsx")    
    return render(request, "index.html", {'data_lines':dir_list})         
        
def fileupload(request):
    form = ImagesForm(request.POST, request.FILES)
    if request.method == 'POST':
        images = request.FILES.getlist('pic')
        for image in images:
            image_ins = Image(pic = image)
            image_ins.save()
        return redirect('index')
    context = {'form': form}
    return render(request, "upload.html", context)
